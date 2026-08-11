
import os, json, uuid, io
from datetime import datetime, date, timedelta
from functools import wraps
from dotenv import load_dotenv
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openai import OpenAI

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "reporthub-secret-2026")

DATA_DIR     = "data"
ROSTER_FILE  = os.path.join(DATA_DIR, "roster.json")
REPORTS_FILE = os.path.join(DATA_DIR, "reports.json")
PROJECTS_FILE= os.path.join(DATA_DIR, "projects.json")
CHANGES_FILE = os.path.join(DATA_DIR, "changes.json")
os.makedirs(DATA_DIR, exist_ok=True)

PROJECT_STATUSES = ["Not Started", "In Progress", "Completed"]
CHANGE_TYPES  = ["Behaviour Rule", "Instruction Update", "New Feature",
                 "Knowledge Base", "Greeting Update", "Bug Fix", "Other"]
CHANGE_STATUSES = ["Live", "Pending Review", "Reverted"]

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

GENERATE_MODES = {
    "standard":   "Write it the way a real person writes a quick work update: plain, direct, human, professional.",
    "fluency":    "Prioritize smooth, natural sentence flow with sentence structure that varies a little bullet to bullet. Still plain, direct, human, professional.",
    "humanize":   "Make it sound like a real person casually recapping their day to their manager: relaxed but professional, contractions are fine, avoid stiff corporate phrasing entirely.",
    "formal":     "Use a formal, polished register suitable for a status report to senior leadership: full professional vocabulary, no contractions, no casual phrasing.",
    "simple":     "Use short, simple sentences and plain everyday words. Avoid jargon and complex phrasing so it reads instantly.",
    "creative":   "Use varied, engaging phrasing and word choice bullet to bullet while staying accurate and professional, avoid sounding templated or repetitive.",
    "expressive": "Use vivid, energetic language that conveys initiative and impact, favor strong active verbs, while staying accurate, truthful, and professional, avoid sounding flat or robotic.",
    "expand":     "Write it the way a real person writes a work update: plain, direct, human, professional. For each bullet, add 2 to 3 sentences of relevant context, likely impact, or reasonable next steps implied by the note, without inventing facts the note doesn't imply.",
    "shorten":    "Write it the way a real person writes a quick work update: plain, direct, human, professional. Strip every bullet down to the essential fact only, no elaboration beyond what the note already says.",
}

# ── Data helpers ───────────────────────────────────────────────────────────────
def load_roster():
    if os.path.exists(ROSTER_FILE):
        with open(ROSTER_FILE, encoding="utf-8") as f: return json.load(f)
    return []

def save_roster(data):
    with open(ROSTER_FILE, "w", encoding="utf-8") as f: json.dump(data, f, indent=2, ensure_ascii=False)

def load_reports():
    if os.path.exists(REPORTS_FILE):
        with open(REPORTS_FILE, encoding="utf-8") as f: return json.load(f)
    return []

def save_reports(data):
    with open(REPORTS_FILE, "w", encoding="utf-8") as f: json.dump(data, f, indent=2, ensure_ascii=False)

def load_projects():
    if os.path.exists(PROJECTS_FILE):
        with open(PROJECTS_FILE, encoding="utf-8") as f: return json.load(f)
    return []

def save_projects(data):
    with open(PROJECTS_FILE, "w", encoding="utf-8") as f: json.dump(data, f, indent=2, ensure_ascii=False)

def load_changes():
    if os.path.exists(CHANGES_FILE):
        with open(CHANGES_FILE, encoding="utf-8") as f: return json.load(f)
    return []

def save_changes(data):
    with open(CHANGES_FILE, "w", encoding="utf-8") as f: json.dump(data, f, indent=2, ensure_ascii=False)

def try_parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None

def project_status_badge(status):
    return {"Completed": "completed", "In Progress": "inprogress",
            "Not Started": "notstarted"}.get(status, "notstarted")

def change_status_badge(status):
    return {"Live": "live", "Pending Review": "pendingreview",
            "Reverted": "reverted"}.get(status, "live")

def today_key():
    return date.today().isoformat()

def pretty_date(key):
    d = datetime.strptime(key, "%Y-%m-%d")
    return d.strftime("%A, %d %B %Y")

def week_dates(monday_str):
    monday = datetime.strptime(monday_str, "%Y-%m-%d").date()
    return [(monday + timedelta(days=i)).isoformat() for i in range(7)]

def get_monday(d=None):
    d = d or date.today()
    return (d - timedelta(days=d.weekday())).isoformat()

def get_recent_reports(name, before_date, limit=None):
    reports = load_reports()
    mine = sorted(
        [r for r in reports
         if r["name"]==name and r["status"]=="submitted" and r["date"] < before_date],
        key=lambda x: x["date"], reverse=True
    )
    return mine if limit is None else mine[:limit]

def today_submission_status(roster, reports):
    tk = today_key()
    submitted_names = {r["name"] for r in reports
                       if r["date"]==tk and r["status"]=="submitted"}
    pending_today   = [m for m in roster if m["name"] not in submitted_names]
    submitted_today = sorted(
        [r for r in reports if r["date"]==tk and r["status"]=="submitted"],
        key=lambda x: [m["name"] for m in roster].index(x["name"])
        if x["name"] in [m["name"] for m in roster] else 999
    )
    return pending_today, submitted_today

# ── Auth ───────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session or session["user"].get("role") != "admin":
            flash("Admin access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        roster = load_roster()
        member = next((m for m in roster if m.get("username","").lower() == username.lower()), None)
        if not member:
            flash("Invalid username or password.", "error")
        elif not member.get("password_hash"):
            flash("This account hasn't been activated yet. Ask your admin to set a password.", "error")
        elif not check_password_hash(member["password_hash"], password):
            flash("Invalid username or password.", "error")
        else:
            session["user"] = {k: v for k, v in member.items() if k != "password_hash"}
            return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    roster   = load_roster()
    reports  = load_reports()
    projects = load_projects()
    changes  = load_changes()
    project_names = {p["id"]: p["name"] for p in projects}

    total = len(projects)
    completed   = sum(1 for p in projects if p["status"]=="Completed")
    in_progress = sum(1 for p in projects if p["status"]=="In Progress")
    not_started = sum(1 for p in projects if p["status"]=="Not Started")
    avg_percent = round(sum(p["percent_complete"] for p in projects)/total) if total else 0

    today = date.today()
    upcoming = []
    for p in projects:
        d = try_parse_date(p["deadline"])
        if d:
            upcoming.append({**p, "deadline_date": d, "overdue": d < today and p["status"]!="Completed"})
    upcoming.sort(key=lambda p: p["deadline_date"])

    recent_changes = sorted(changes, key=lambda c: c["added_at"], reverse=True)[:15]
    for c in recent_changes:
        c["project_name"] = project_names.get(c["project_id"], "Unknown project")

    pending_today, submitted_today = today_submission_status(roster, reports)

    return render_template("dashboard.html",
        projects=projects, total=total, completed=completed, in_progress=in_progress,
        not_started=not_started, avg_percent=avg_percent, upcoming=upcoming,
        recent_changes=recent_changes, pending_today=pending_today,
        submitted_today=submitted_today, project_status_badge=project_status_badge,
        change_status_badge=change_status_badge, today_pretty=pretty_date(today_key()))

@app.route("/today", methods=["GET", "POST"])
@login_required
def today():
    user    = session["user"]
    reports = load_reports()
    tk      = today_key()
    existing = next((r for r in reports if r["name"]==user["name"] and r["date"]==tk), None)
    return render_template("today.html", user=user, existing=existing,
                           today_pretty=pretty_date(tk))

@app.route("/save-report", methods=["POST"])
@login_required
def save_report():
    user     = session["user"]
    reports  = load_reports()
    tk       = today_key()
    raw_work    = request.form.get("raw_work","").strip()
    raw_pending = request.form.get("raw_pending","").strip()
    generated   = request.form.get("generated","").strip()
    status      = request.form.get("status","draft")
    idx = next((i for i,r in enumerate(reports)
                if r["name"]==user["name"] and r["date"]==tk), None)
    entry = {
        "id":           reports[idx]["id"] if idx is not None else str(uuid.uuid4())[:8],
        "name":         user["name"],
        "date":         tk,
        "raw_work":     raw_work,
        "raw_pending":  raw_pending,
        "generated":    generated,
        "status":       status,
        "submitted_at": datetime.now().isoformat() if status=="submitted" else None,
    }
    if idx is not None: reports[idx] = entry
    else:               reports.append(entry)
    save_reports(reports)
    flash("Submitted!" if status=="submitted" else "Saved as draft.", "success")
    return redirect(url_for("today"))

@app.route("/generate", methods=["POST"])
@login_required
def generate():
    user        = session["user"]
    raw_work    = request.json.get("raw_work","").strip()
    raw_pending = request.json.get("raw_pending","").strip()
    mode        = request.json.get("mode","standard").strip().lower()
    if mode not in GENERATE_MODES:
        mode = "standard"
    if not raw_work:
        return jsonify({"error": "Add a note on what you worked on first."}), 400
    if not openai_client:
        return jsonify({"error": "API key not configured."}), 500

    past = get_recent_reports(user["name"], today_key())
    past_context = "\n\n".join(
        f"{pretty_date(r['date'])}: {r.get('generated') or r.get('raw_work','')}" for r in past
    ) or "none"

    if mode == "expand":
        length_rule = ('Each bullet starts with "- ", first person, past tense. Use as many bullets as the '
                        "notes actually support, usually 3-6.")
    elif mode == "shorten":
        length_rule = ('Each bullet starts with "- ", first person, past tense, as short as possible while '
                        "still being clear, ideally under 12 words. Combine closely related notes into one "
                        "bullet instead of splitting them.")
    else:
        length_rule = ('Each bullet starts with "- ", is one line, first person, past tense. Use as many '
                        "bullets as the notes actually support, usually 3-6, don't pad with filler.")

    prompt = f"""You are turning a team member's rough end-of-day notes into a daily work update for their manager. {GENERATE_MODES[mode]} No AI phrases like 'I successfully accomplished', 'delved into', 'leveraged', or corporate filler.

Format: a short bullet list. {length_rule}

The notes are often very brief, sometimes just a few words per task, because the person is jotting them down quickly at the end of the day. Turn each note into a fuller, clearer bullet using reasonable professional detail and context, but only elaborate on what's actually implied by the note, never invent tasks, names, or outcomes that aren't there. Use the person's full report history below to understand shorthand references, recurring names, cases, or threads, especially when today's notes are vague, abbreviated, or unclear on their own, do not copy or repeat this history's content into today's bullets.

Full report history from {user['name']}, oldest context, do not repeat or summarize these:
{past_context}

Date: {pretty_date(today_key())}

Notes on what was done today:
{raw_work}

Notes on what's pending:
{raw_pending or 'none'}

If there is anything pending, add it as one extra bullet at the end (e.g. "Still waiting on..." / "Need to follow up on..."). Never use an em dash, use a period, comma, or colon instead. Reply with ONLY the finished bullet list, no preamble, no headers, no quotation marks, no markdown besides the leading "- " on each line."""
    try:
        max_tokens = {"expand": 900, "shorten": 250}.get(mode, 500)
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}],
        )
        text = (resp.choices[0].message.content or "").strip()
        if not text:
            return jsonify({"error": "The model returned an empty response, try again."}), 502
        return jsonify({"text": text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/transcribe", methods=["POST"])
@login_required
def transcribe():
    if not openai_client:
        return jsonify({"error": "API key not configured."}), 500
    audio_file = request.files.get("audio")
    if not audio_file or audio_file.filename == "":
        return jsonify({"error": "No audio received."}), 400
    try:
        audio_buf = io.BytesIO(audio_file.read())
        audio_buf.name = "note.webm"
        resp = openai_client.audio.transcriptions.create(
            model="whisper-1",
            file=audio_buf,
        )
        text = (resp.text or "").strip()
        if not text:
            return jsonify({"error": "Couldn't make out any speech, try again."}), 502
        return jsonify({"text": text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/reports/<report_id>/edit", methods=["GET", "POST"])
@login_required
def report_edit(report_id):
    user     = session["user"]
    is_admin = user.get("role") == "admin"
    reports  = load_reports()
    idx = next((i for i,r in enumerate(reports) if r["id"]==report_id), None)
    if idx is None:
        flash("Report not found.", "error")
        return redirect(url_for("my_reports"))
    is_own = reports[idx]["name"] == user["name"]
    if not is_own and not is_admin:
        flash("You can only edit your own reports.", "error")
        return redirect(url_for("my_reports"))
    back_endpoint = "my_reports" if is_own else "admin_reports"
    back_args = {} if is_own else {"member": reports[idx]["name"]}
    if request.method == "POST":
        reports[idx].update({
            "raw_work":    request.form.get("raw_work","").strip(),
            "raw_pending": request.form.get("raw_pending","").strip(),
            "generated":   request.form.get("generated","").strip(),
        })
        if not is_own:
            reports[idx]["edited_by"] = user["name"]
            reports[idx]["edited_at"] = datetime.now().isoformat()
        save_reports(reports)
        flash("Report updated.", "success")
        return redirect(url_for(back_endpoint, **back_args))
    return render_template("report_edit.html", report=reports[idx], pretty_date=pretty_date,
                           is_own=is_own, back_endpoint=back_endpoint, back_args=back_args)

@app.route("/team/reports")
@login_required
@admin_required
def admin_reports():
    roster  = load_roster()
    reports = load_reports()
    member  = request.args.get("member", roster[0]["name"] if roster else "")
    mine    = sorted([r for r in reports if r["name"]==member],
                     key=lambda x: x["date"], reverse=True)
    return render_template("admin_reports.html", roster=roster, member=member,
                           reports=mine, pretty_date=pretty_date, today_key=today_key())

@app.route("/team/log-report", methods=["POST"])
@login_required
@admin_required
def log_report():
    user        = session["user"]
    roster      = load_roster()
    member_name = request.form.get("member", "").strip()
    report_date = request.form.get("date", "").strip()
    raw_work    = request.form.get("raw_work", "").strip()
    raw_pending = request.form.get("raw_pending", "").strip()

    if not any(m["name"] == member_name for m in roster):
        flash("Member not found.", "error")
        return redirect(url_for("admin_reports"))

    parsed = try_parse_date(report_date)
    if not parsed:
        flash("Enter a valid date.", "error")
        return redirect(url_for("admin_reports", member=member_name))
    if parsed > date.today():
        flash("Can't log a report for a future date.", "error")
        return redirect(url_for("admin_reports", member=member_name))
    if not raw_work:
        flash("Add the report text before logging it.", "error")
        return redirect(url_for("admin_reports", member=member_name))

    reports = load_reports()
    existing = next((r for r in reports if r["name"]==member_name and r["date"]==report_date), None)
    if existing:
        flash(f"{member_name} already has a report for {pretty_date(report_date)}. Edit it instead.", "error")
        return redirect(url_for("report_edit", report_id=existing["id"]))

    now = datetime.now().isoformat()
    reports.append({
        "id":           str(uuid.uuid4())[:8],
        "name":         member_name,
        "date":         report_date,
        "raw_work":     raw_work,
        "raw_pending":  raw_pending,
        "generated":    "",
        "status":       "submitted",
        "submitted_at": now,
        "logged_by":    user["name"],
        "logged_at":    now,
    })
    save_reports(reports)
    flash(f"Logged {pretty_date(report_date)}'s report for {member_name}.", "success")
    return redirect(url_for("admin_reports", member=member_name))

@app.route("/my-reports")
@login_required
def my_reports():
    user    = session["user"]
    reports = load_reports()
    mine    = sorted([r for r in reports if r["name"]==user["name"]],
                     key=lambda x: x["date"], reverse=True)
    return render_template("my_reports.html", user=user, reports=mine,
                           pretty_date=pretty_date)

@app.route("/download-my-reports")
@login_required
def download_my_reports():
    user    = session["user"]
    reports = load_reports()
    mine    = sorted([r for r in reports
                      if r["name"]==user["name"] and r["status"]=="submitted"],
                     key=lambda x: x["date"])
    wb = Workbook()
    ws = wb.active
    ws.title = "My Reports"
    ws.append(["Date","Report","Status"])
    for r in mine:
        ws.append([pretty_date(r["date"]), r.get("generated") or r.get("raw_work",""), r["status"]])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 12
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"{user['name']}_reports.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/projects")
@login_required
def projects_list():
    projects = load_projects()
    status_filter = request.args.get("status", "")
    if status_filter:
        projects = [p for p in projects if p["status"]==status_filter]
    return render_template("projects.html", projects=projects, status_filter=status_filter,
                           statuses=PROJECT_STATUSES, project_status_badge=project_status_badge)

@app.route("/projects/new", methods=["GET", "POST"])
@login_required
@admin_required
def project_new():
    if request.method == "POST":
        projects = load_projects()
        project_id = request.form.get("id", "").strip()
        if not project_id:
            flash("Project ID is required.", "error")
            return redirect(url_for("project_new"))
        if any(p["id"]==project_id for p in projects):
            flash(f"Project ID {project_id} already exists.", "error")
            return redirect(url_for("project_new"))
        now = datetime.now().isoformat()
        projects.append({
            "id": project_id,
            "name": request.form.get("name","").strip(),
            "description": request.form.get("description","").strip(),
            "owner": request.form.get("owner","").strip(),
            "progress_summary": request.form.get("progress_summary","").strip(),
            "status": request.form.get("status","Not Started"),
            "percent_complete": int(request.form.get("percent_complete") or 0),
            "start_date": request.form.get("start_date","").strip(),
            "deadline": request.form.get("deadline","").strip(),
            "risks": request.form.get("risks","").strip(),
            "next_action": request.form.get("next_action","").strip(),
            "kpi": request.form.get("kpi","").strip(),
            "created_at": now,
            "updated_at": now,
        })
        save_projects(projects)
        flash(f"Project {project_id} added.", "success")
        return redirect(url_for("projects_list"))
    return render_template("project_form.html", project=None, statuses=PROJECT_STATUSES)

@app.route("/projects/<project_id>")
@login_required
def project_detail(project_id):
    projects = load_projects()
    project = next((p for p in projects if p["id"]==project_id), None)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("projects_list"))
    changes = sorted(
        [c for c in load_changes() if c["project_id"]==project_id],
        key=lambda c: c["added_at"], reverse=True
    )
    return render_template("project_detail.html", project=project, changes=changes,
                           project_status_badge=project_status_badge,
                           change_status_badge=change_status_badge)

@app.route("/projects/<project_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def project_edit(project_id):
    projects = load_projects()
    idx = next((i for i,p in enumerate(projects) if p["id"]==project_id), None)
    if idx is None:
        flash("Project not found.", "error")
        return redirect(url_for("projects_list"))
    if request.method == "POST":
        p = projects[idx]
        p.update({
            "name": request.form.get("name","").strip(),
            "description": request.form.get("description","").strip(),
            "owner": request.form.get("owner","").strip(),
            "progress_summary": request.form.get("progress_summary","").strip(),
            "status": request.form.get("status","Not Started"),
            "percent_complete": int(request.form.get("percent_complete") or 0),
            "start_date": request.form.get("start_date","").strip(),
            "deadline": request.form.get("deadline","").strip(),
            "risks": request.form.get("risks","").strip(),
            "next_action": request.form.get("next_action","").strip(),
            "kpi": request.form.get("kpi","").strip(),
            "updated_at": datetime.now().isoformat(),
        })
        save_projects(projects)
        flash(f"Project {project_id} updated.", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    return render_template("project_form.html", project=projects[idx], statuses=PROJECT_STATUSES)

@app.route("/changes")
@login_required
def changes_list():
    changes  = load_changes()
    projects = load_projects()
    project_names = {p["id"]: p["name"] for p in projects}
    project_filter = request.args.get("project", "")
    type_filter    = request.args.get("change_type", "")
    if project_filter:
        changes = [c for c in changes if c["project_id"]==project_filter]
    if type_filter:
        changes = [c for c in changes if c["change_type"]==type_filter]
    changes = sorted(changes, key=lambda c: c["added_at"], reverse=True)
    for c in changes:
        c["project_name"] = project_names.get(c["project_id"], "Unknown project")
    return render_template("changes.html", changes=changes, projects=projects,
                           change_types=CHANGE_TYPES, project_filter=project_filter,
                           type_filter=type_filter, change_status_badge=change_status_badge)

@app.route("/changes/new", methods=["GET", "POST"])
@login_required
def change_new():
    projects = load_projects()
    if request.method == "POST":
        user = session["user"]
        changes = load_changes()
        now = datetime.now().isoformat()
        changes.append({
            "id": str(uuid.uuid4())[:8],
            "project_id": request.form.get("project_id",""),
            "change_type": request.form.get("change_type", CHANGE_TYPES[0]),
            "summary": request.form.get("summary","").strip(),
            "description": request.form.get("description","").strip(),
            "status": request.form.get("status", CHANGE_STATUSES[0]),
            "added_by": user["name"],
            "added_at": now,
            "edited_by": None,
            "edited_at": None,
        })
        save_changes(changes)
        flash("Change logged.", "success")
        return redirect(url_for("changes_list"))
    return render_template("change_form.html", change=None, projects=projects,
                           change_types=CHANGE_TYPES, change_statuses=CHANGE_STATUSES,
                           preselect_project=request.args.get("project", ""))

@app.route("/changes/<change_id>/edit", methods=["GET", "POST"])
@login_required
def change_edit(change_id):
    changes = load_changes()
    projects = load_projects()
    idx = next((i for i,c in enumerate(changes) if c["id"]==change_id), None)
    if idx is None:
        flash("Change not found.", "error")
        return redirect(url_for("changes_list"))
    if request.method == "POST":
        user = session["user"]
        c = changes[idx]
        c.update({
            "project_id": request.form.get("project_id",""),
            "change_type": request.form.get("change_type", CHANGE_TYPES[0]),
            "summary": request.form.get("summary","").strip(),
            "description": request.form.get("description","").strip(),
            "status": request.form.get("status", CHANGE_STATUSES[0]),
            "edited_by": user["name"],
            "edited_at": datetime.now().isoformat(),
        })
        save_changes(changes)
        flash("Change updated.", "success")
        return redirect(url_for("changes_list"))
    return render_template("change_form.html", change=changes[idx], projects=projects,
                           change_types=CHANGE_TYPES, change_statuses=CHANGE_STATUSES)

@app.route("/team")
@login_required
@admin_required
def team():
    roster  = load_roster()
    reports = load_reports()
    tk      = today_key()
    monday  = request.args.get("monday", get_monday())
    pending_today, submitted_today = today_submission_status(roster, reports)
    return render_template("team.html", user=session["user"], roster=roster,
                           pending_today=pending_today, submitted_today=submitted_today,
                           monday=monday, pretty_date=pretty_date,
                           today_pretty=pretty_date(tk))

@app.route("/download-week")
@login_required
@admin_required
def download_week():
    roster  = load_roster()
    reports = load_reports()
    monday  = request.args.get("monday", get_monday())
    days    = week_dates(monday)
    names   = [m["name"] for m in roster]
    wb  = Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    hf  = PatternFill("solid", fgColor="1C3557")
    bd  = Border(
        left=Side(style="thin",color="CCCCCC"),
        right=Side(style="thin",color="CCCCCC"),
        top=Side(style="thin",color="CCCCCC"),
        bottom=Side(style="thin",color="CCCCCC")
    )
    first = True
    for day in days:
        day_reports = {r["name"]: r for r in reports
                       if r["date"]==day and r["status"]=="submitted"}
        if not day_reports: continue
        ws = wb.active if first else wb.create_sheet()
        ws.title = pretty_date(day)[:28]
        first = False
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 85
        for col, title in enumerate(["Name","Daily Report"],1):
            c = ws.cell(row=1, column=col, value=title)
            c.font = hdr; c.fill = hf
            c.alignment = Alignment(vertical="center")
            c.border = bd
        ws.row_dimensions[1].height = 22
        row = 2
        for name in names:
            if name not in day_reports: continue
            r = day_reports[name]
            text = r.get("generated") or r.get("raw_work","")
            c1 = ws.cell(row=row, column=1, value=name)
            c2 = ws.cell(row=row, column=2, value=text)
            for c in [c1,c2]:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = bd
            ws.row_dimensions[row].height = max(40, len(text)//6)
            row += 1
    if first:
        ws = wb.active; ws.title = "No data"
        ws.cell(row=1,column=1,value="No submitted reports for this week.")
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"weekly_report_{monday}.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def styled_export_sheet(ws, headers, col_widths):
    hdr = Font(bold=True, color="FFFFFF")
    hf  = PatternFill("solid", fgColor="1C3557")
    bd  = Border(
        left=Side(style="thin",color="CCCCCC"),
        right=Side(style="thin",color="CCCCCC"),
        top=Side(style="thin",color="CCCCCC"),
        bottom=Side(style="thin",color="CCCCCC")
    )
    for col, title in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = hdr; c.fill = hf
        c.alignment = Alignment(vertical="center")
        c.border = bd
        ws.column_dimensions[chr(64+col)].width = col_widths[col-1]
    ws.row_dimensions[1].height = 22
    return bd

@app.route("/download-projects")
@login_required
@admin_required
def download_projects():
    projects = load_projects()
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = ["Project ID","Project Name","Description","Owner","What Has Been Done",
               "Status","Percent Complete","Start Date","Deadline","Risks",
               "Next Action","KPI"]
    widths = [12,26,45,18,45,14,10,14,14,40,40,30]
    bd = styled_export_sheet(ws, headers, widths)
    row = 2
    for p in projects:
        values = [p["id"], p["name"], p["description"], p["owner"], p["progress_summary"],
                  p["status"], p["percent_complete"], p["start_date"], p["deadline"],
                  p["risks"], p["next_action"], p["kpi"]]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = bd
        ws.row_dimensions[row].height = 40
        row += 1
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"projects_{today_key()}.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download-changes")
@login_required
@admin_required
def download_changes():
    changes  = sorted(load_changes(), key=lambda c: c["added_at"], reverse=True)
    projects = load_projects()
    project_names = {p["id"]: p["name"] for p in projects}
    wb = Workbook()
    ws = wb.active
    ws.title = "Changes"
    headers = ["Date","Time","Project","Change Type","Summary","Description",
               "Status","Added By"]
    widths = [14,10,22,18,30,45,14,16]
    bd = styled_export_sheet(ws, headers, widths)
    row = 2
    for c in changes:
        added = datetime.fromisoformat(c["added_at"])
        values = [added.strftime("%d %B %Y"), added.strftime("%H:%M"),
                  project_names.get(c["project_id"], "Unknown project"),
                  c["change_type"], c["summary"], c["description"],
                  c["status"], c["added_by"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = bd
        ws.row_dimensions[row].height = 40
        row += 1
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"changes_{today_key()}.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/add-member", methods=["POST"])
@login_required
@admin_required
def add_member():
    name     = request.form.get("name","").strip()
    username = request.form.get("username","").strip()
    password = request.form.get("password","")
    role     = request.form.get("role","member")
    if not name or not username or not password:
        flash("Name, username and password are all required.", "error")
        return redirect(url_for("team"))
    roster = load_roster()
    if any(m["name"].lower()==name.lower() for m in roster):
        flash(f"{name} is already on the roster.", "error")
        return redirect(url_for("team"))
    if any(m.get("username","").lower()==username.lower() for m in roster):
        flash(f"Username {username} is already taken.", "error")
        return redirect(url_for("team"))
    roster.append({
        "id": str(uuid.uuid4())[:8],
        "name": name,
        "username": username.lower(),
        "password_hash": generate_password_hash(password),
        "role": role,
    })
    save_roster(roster)
    flash(f"{name}'s account created.", "success")
    return redirect(url_for("team"))

@app.route("/update-member", methods=["POST"])
@login_required
@admin_required
def update_member():
    member_id    = request.form.get("id")
    new_name     = request.form.get("name","").strip()
    new_username = request.form.get("username","").strip()
    new_password = request.form.get("password","")
    new_role     = request.form.get("role", "member")
    if new_role not in ("member", "admin"):
        new_role = "member"
    if not new_name or not new_username:
        flash("Name and username can't be empty.", "error")
        return redirect(url_for("team"))
    roster = load_roster()
    member = next((m for m in roster if m["id"] == member_id), None)
    if not member:
        flash("Member not found.", "error")
        return redirect(url_for("team"))
    if any(m["id"] != member_id and m["name"].lower() == new_name.lower() for m in roster):
        flash(f"{new_name} is already on the roster.", "error")
        return redirect(url_for("team"))
    if any(m["id"] != member_id and m.get("username","").lower() == new_username.lower() for m in roster):
        flash(f"Username {new_username} is already taken.", "error")
        return redirect(url_for("team"))
    if member["role"] == "admin" and new_role != "admin":
        admin_count = sum(1 for m in roster if m["role"] == "admin")
        if admin_count <= 1:
            flash("Can't remove the last admin. Promote someone else first.", "error")
            return redirect(url_for("team"))

    old_name = member["name"]
    member["name"] = new_name
    member["username"] = new_username.lower()
    member["role"] = new_role
    if new_password:
        member["password_hash"] = generate_password_hash(new_password)
    save_roster(roster)

    if old_name != new_name:
        reports = load_reports()
        for r in reports:
            if r["name"] == old_name:
                r["name"] = new_name
        save_reports(reports)

    if session["user"]["id"] == member_id:
        session["user"]["name"] = new_name
        session["user"]["role"] = new_role

    flash(f"{new_name}'s details updated.", "success")
    return redirect(url_for("team"))

@app.route("/remove-member", methods=["POST"])
@login_required
@admin_required
def remove_member():
    member_id = request.form.get("id")
    roster = load_roster()
    roster = [m for m in roster if m["id"] != member_id]
    save_roster(roster)
    flash("Member removed.", "success")
    return redirect(url_for("team"))

# Custom Jinja2 filters
@app.template_filter('enumerate')
def enumerate_filter(iterable):
    return enumerate(iterable)

@app.context_processor
def inject_user():
    return {"current_user": session.get("user")}

if __name__ == "__main__":
    app.run(debug=True, port=5001)
