"""One-time import of Projects data from R&D Project Track.xlsx into data/projects.json.

Run with: py import_projects.py
Re-running is safe: it refuses to overwrite an existing non-empty
data/projects.json unless called with --force.
"""
import argparse
import json
import os
import re
from datetime import datetime, date

import openpyxl

XLSX_FILE    = "R&D Project Track.xlsx"
PROJECTS_FILE = os.path.join("data", "projects.json")


def strip_em_dash(text):
    if not isinstance(text, str):
        return text
    text = text.replace(" — ", ": ").replace("—", ":")
    return text


def clean_field(value):
    if isinstance(value, str):
        return strip_em_dash(value.strip())
    return value or ""


def normalize_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    try:
        return datetime.strptime(text, "%d/%m/%Y").date().isoformat()
    except ValueError:
        return text


def normalize_percent(value):
    if isinstance(value, (int, float)):
        return round(value * 100)
    return 0


def find_projects_sheet(wb):
    for name in wb.sheetnames:
        if "project" in name.lower():
            return wb[name]
    raise ValueError("No sheet with 'Projects' in its name was found in the workbook.")


def import_projects(wb=None):
    """Parse the Projects sheet into a list of project dicts. Pass an already-open
    `wb` (e.g. from an uploaded file) to reuse this parsing without touching disk;
    omitting it falls back to reading XLSX_FILE, for the CLI one-time-import use
    below."""
    if wb is None:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = find_projects_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))

    # Skip the two title rows and the column-header row.
    data_rows = rows[3:]

    projects = []
    now = datetime.now().isoformat()
    for row in data_rows:
        project_id = row[0]
        if not project_id or not str(project_id).strip():
            continue
        projects.append({
            "id":               str(project_id).strip(),
            "name":             clean_field(row[1]),
            "description":      clean_field(row[2]),
            "owner":            clean_field(row[3]),
            "progress_summary": clean_field(row[4]),
            "status":           clean_field(row[5]) or "Not Started",
            "percent_complete": normalize_percent(row[6]),
            "start_date":       strip_em_dash(normalize_date(row[7])),
            "deadline":         strip_em_dash(normalize_date(row[8])),
            "risks":            clean_field(row[9]),
            "next_action":      clean_field(row[10]),
            "kpi":              clean_field(row[11]),
            "created_at":       now,
            "updated_at":       now,
        })
    return projects


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true",
                         help="Overwrite an existing non-empty data/projects.json")
    args = parser.parse_args()

    if os.path.exists(PROJECTS_FILE):
        with open(PROJECTS_FILE, encoding="utf-8") as f:
            existing = json.load(f)
        if existing and not args.force:
            print(f"{PROJECTS_FILE} already has {len(existing)} project(s). "
                  "Re-run with --force to overwrite.")
            return

    try:
        projects = import_projects()
    except ValueError as e:
        raise SystemExit(str(e))
    os.makedirs("data", exist_ok=True)
    with open(PROJECTS_FILE, "w", encoding="utf-8") as f:
        json.dump(projects, f, indent=2, ensure_ascii=False)
    print(f"Imported {len(projects)} project(s) into {PROJECTS_FILE}.")


if __name__ == "__main__":
    main()
